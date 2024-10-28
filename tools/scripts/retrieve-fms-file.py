#!/usr/bin/env python3

import os
import sys
import boto3
import argparse
import requests
import urllib.request
import simplejson as json
from io import BytesIO
from dotenv import dotenv_values
from openpyxl import load_workbook

config = dotenv_values("../../local.env")
s3_client = boto3.client('s3')
pcf_filename = "pcf_apm_entity_ids.json"


def get_user_inputs():
    parser = argparse.ArgumentParser()
    parser.add_argument('-au', '--auth-url', required=True, type=str,
                        help='QPP Auth token retrieval url. Example: https://imp.qpp.cms.gov/api/auth/oauth/token')
    parser.add_argument('-fu', '--fms-url', required=True, type=str,
                        help='FMS Base url. Example: https://impl.ar.qpp.internal/dataservices')
    parser.add_argument('-t', '--fms-token', required=True, type=str,
                        help='QPP Auth client assertion token to retrieve the FMS S2S token')
    parser.add_argument('-p', '--fms-path', required=True, type=str,
                        help='FMS path with file name and extension. Example: /folder/file.xlsx')
    args = parser.parse_args()

    return args


def download_from_fms(auth_url, fms_url, fms_token, fms_path, filename):
    d = {'client_assertion': fms_token,
         'client_assertion_type': 'urn:ietf:params:oauth:client-assertion-type:jwt-bearer',
         'grant_type': 'client_credentials',
         'scope': 'analyticsAndReporting'
         }
    # print('starting s2s token retrieval request from qpp auth')
    get_s2s_token = requests.post(
        url=auth_url,
        data=d,
        headers={
            'Content-Type': 'application/x-www-form-urlencoded',
            'Accept': 'application/vnd.qpp.cms.gov.v2+json'
        }
    )
    s2s_token = get_s2s_token.json()["data"]["token"]
    # print('starting download from fms for file - ' + filename)
    get_download_url = requests.post(
        url=fms_url + '/get-file',
        json={"path": fms_path + filename},
        verify=False,
        headers={
            'Accept': 'application/vnd.qpp.cms.gov.v2+json',
            'Authorization': 'Bearer ' + s2s_token
        }
    )
    download_url = get_download_url.json()['presigned_url']
    urllib.request.urlretrieve(download_url, filename)


def process_file(filename):
    print('processing file')
    wb = load_workbook(filename)
    sh = wb['2024_Practices']
    data_list = []
    for row in sh.iter_rows(sh.min_row + 1, sh.max_row):
        data_list.append(row[0].value)
    json_data = json.dumps(data_list)
    return str(json_data).replace(" ", "")

def update_local_repo(data):
    # print('writing ' + pcf_filename + ' to local repository')
    with open('../../converter/src/main/resources/' + pcf_filename, 'w') as f:
        f.write(data)


def upload_to_s3(data):
    # print('starting to upload file to s3 bucket - ' + pcf_filename)
    upload_status = s3_client.put_object(
        Bucket=config.get('s3_bucket'),
        Key=pcf_filename,
        Body=data,
        ContentType='application/json',
        ServerSideEncryption='aws:kms'
    )
    print(upload_status)

def delete_file(filename):
    if os.path.exists("./" + filename):
        os.remove("./" + filename)
        print("File " + filename + " has been processed and removed successfully!")
    else:
        print("Can not process or delete file " + filename + ", as it doesn't exists")

def main():
    try:
        # args = get_user_inputs()
        # s3_url = download_from_fms(args.auth_url, args.fms_url, args.fms_token, args.fms_path)
        filename = config.get('filename')
        download_from_fms(config.get('auth_url'), config.get('fms_url'), config.get('fms_token'),
                                            config.get('fms_path'), filename)
        processed_data = process_file(filename)
        update_local_repo(processed_data)
        # upload_to_s3(processed_data)
        delete_file(filename)
    except Exception as err:
        print(f"Unexpected Error. {err = }, {type(err) = }")
        sys.exit(1)


if __name__ == '__main__':
    main()
