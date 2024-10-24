#! /usr/bin/env python
from openpyxl import load_workbook
from collections import OrderedDict
import simplejson as json
import sys

def main(argv):
	# Open the workbook and select the second sheet
	wb = load_workbook(filename = argv[0])
	sh = wb['2024_Providers']
	data_list = []
	for row in sh.iter_rows(sh.min_row+1, sh.max_row):
		data = OrderedDict()
		data['npi'] = row[10].value
		data['tin'] = row[11].value
		data['apm_entity_id'] = row[0].value
		data_list.append(data)
	j = json.dumps(data_list)
	# Write to file (cpc-validation-file.json for prod & dev pre. testCpcPlusValidationFile.json for testing)
	with open(argv[1], 'w') as f:
		f.write(j)

if __name__ == "__main__":
	main(sys.argv[1:])